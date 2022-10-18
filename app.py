# -*- encoding: utf-8 -*-
from flask import Flask
# from config import DevConfig
from datetime import datetime
import random
import time
from core.crawler import Crawler
from openpyxl import Workbook
from core.browsers import SLEEP_SECOND


app = Flask(__name__)
# app.config.from_object(DevConfig)


@app.route("/")
def home():
    return "Hello Flask {}".format(random.randint(1000, 999999))


@app.route("/test/")
def test():
    return "test 789 - {}".format(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


def run_crawler():
    # areas = [
    #     ["Africa", "Africa", "Nigeria", "160", "nigeria"],
    #     ["Africa", "Africa", "South Africa 南非", "112", "south-africa"],
    #     ["Africa", "Africa", "Morocco", "159", "morocco"],
    #     ["Africa", "Africa", "Kenya", "247", "kenya"],
    #     ["Asia", "East Asia", "China 中國", "117", "china"],
    #     ["Asia", "East Asia", "Japan 日本", "121", "japan"],
    #     ["Asia", "East Asia", "South Korea", "125", "south-korea"],
    #     ["Asia", "East Asia", "Hong Kong 香港", "118", "hong-kong"],
    #     ["Asia", "SEA", "Indonesia 印尼", "120", "indonesia"],
    #     ["Asia", "SEA", "Thailand 泰國", "126", "thailand"],
    #     ["Asia", "SEA", "Vietnam 越南", "127", "vietnam"],
    #     ["Asia", "South Asia", "India 印度", "119", "india"],
    #     ["Asia", "South Asia", "Pakistan 巴基斯坦", "294", "pakistan"],
    #     ["Asia", "West Asia", "Saudi Arabia 沙烏地阿拉伯", "110", "saudi-arabia"],
    #     ["Australia & Oceania", "Australia & Oceania", "Australia 澳洲", "107", "australia"],
    #     ["Australia & Oceania", "Australia & Oceania", "New Zealand", "161", "new-zealand"],
    #     ["Europe", "Central & West Europe", "Germany 德國", "137", "germany"],
    #     ["Europe", "Central & West Europe", "UK 英國", "156", "united-kingdom"],
    #     ["Europe", "Central & West Europe", "France 法國", "136", "france"],
    #     ["Europe", "Central & West Europe", "Poland 波蘭", "146", "poland"],
    #     ["Europe", "Central & West Europe", "Netherlands 荷蘭", "144", "netherlands"],
    #     ["Europe", "Central & West Europe", "Switzerland 瑞士", "155", "switzerland"],
    #     ["Europe", "Central & West Europe", "Belgium 比利時", "129", "belgium"],
    #     ["Europe", "Central & West Europe", "Austria 奧地利", "128", "austria"],
    #     ["Europe", "Central & West Europe", "Czechia", "132", "czechia"],
    #     ["Europe", "Central & West Europe", "Ireland", "140", "ireland"],
    #     ["Europe", "Eastern Europe", "Russia 俄羅斯", "149", "russia"],
    #     ["Europe", "Northern Europe", "Sweden 瑞典", "154", "sweden"],
    #     ["Europe", "Northern Europe", "Norway", "145", "norway"],
    #     ["Europe", "Northern Europe", "Denmark", "133", "denmark"],
    #     ["Europe", "Northern Europe", "Finland", "135", "finland"],
    #     ["Europe", "Southern Europe", "Italy 義大利", "141", "italy"],
    #     ["Europe", "Southern Europe", "Spain 西班牙", "153", "spain"],
    #     ["Europe", "Southern Europe", "Turkey 土耳其", "113", "turkey"],
    #     ["Europe", "Southern Europe", "Portugal", "147", "portugal"],
    #     ["Europe", "Southern Europe", "Greece 希臘", "138", "greece"],
    #     ["North America", "North America", "USA 美國", "109", "united-states"],
    #     ["North America", "North America", "Mexico 墨西哥", "116", "mexico"],
    #     ["North America", "North America", "Canada 加拿大", "108", "canada"],
    #     ["South America", "South America", "Brazil 巴西", "115", "brazil"],
    #     ["South America", "South America", "Argentina 阿根廷", "114", "argentina"],
    #     ["South America", "South America", "Colombia 哥倫比亞", "158", "colombia"],
    #     ["South America", "South America", "Chile", "157", "chile"],
    #     ["Worldwide", "Worldwide", "", "100", "worldwide"]
    # ]

    data = {
        "prefix_label": ["國家", ""],
        "areas": [
            ["Worldwide", "全球", "100", "worldwide"],
            ["HongKong", "香港", "118", "hong-kong"],
            ["China", "中國", "117", "china"],
            ["USA", "美國", "109", "united-states"],
            ["India", "印度", "119", "india"],
            ["Japan", "日本", "121", "japan"],
            ["Brazil", "巴西", "115", "brazil"],
            ["Germany", "德國", "137", "germany"],
            ["UK", "英國", "156", "united-kingdom"],
            ["France", "法國", "136", "france"],
            ["Russia", "俄羅斯", "149", "russia"],
            ["Mexico", "墨西哥", "116", "mexico"],
            ["Indonesia", "印尼", "120", "indonesiaa"],
            ["South Korea", "南韓", "125", "south-korea"],
            ["Italy", "義大利", "141", "italy"],
            ["Canada", "加拿大", "108", "canada"],
            ["Spain", "西班牙", "153", "spain"],
            ["Australia", "澳洲", "107", "australia"],
            ["Turkey", "土耳其", "113", "turkey"],
            ["Argentina", "阿根廷", "114", "argentina"],
            ["Saudi Arabia", "沙烏地阿拉伯", "110", "saudi-arabia"],
            ["Poland", "波蘭", "146", "poland"],
            ["Netherlands", "荷蘭", "144", "netherlands"],
            ["Thailand", "泰國", "126", "thailand"],
            ["Vietnam", "越南", "127", "vietnam"],
            ["Colombia", "哥倫比亞", "158", "colombia"],
            ["Switzerland", "瑞士", "155", "switzerland"],
            ["Sweden", "瑞典", "154", "sweden"],
            ["Belgium", "比利時", "129", "belgium"]
        ],
        "tab": {
            "食品類": [
                {
                    "title": "Consumer Market(Non-Carbonated Soft Drinks)",
                    "r_type": "20020200",
                    "i_type": "non-carbonated-soft-drinks",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Roast Coffee)",
                    "r_type": "30010100",
                    "i_type": "roast-coffee",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Instant Coffee)",
                    "r_type": "30010200",
                    "i_type": "instant-coffee",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Tea)",
                    "r_type": "30020000",
                    "i_type": "tea",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Processed & Frozen Fruits)",
                    "r_type": "40040200",
                    "i_type": "processed-frozen-fruits",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Nuts)",
                    "r_type": "40110300",
                    "i_type": "nuts",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Spices & Culinary Herbs)",
                    "r_type": "40070300",
                    "i_type": "spices-culinary-herbs",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Snack Food)",
                    "r_type": "40110000",
                    "i_type": "snack-food",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Baby Food)",
                    "r_type": "40120000",
                    "i_type": "baby-food",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Pet Food)",
                    "r_type": "40130000",
                    "i_type": "pet-food",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Digital Market(Food & Beverage) ",
                    "r_type": "253",
                    "i_type": "food-beverages",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                }
            ],
            "服飾精品類": [
                {
                    "title": "Digital Market(Fashion)",
                    "r_type": "244",
                    "i_type": "fashion",
                    "labels": ["Revenue", "yoy", "CAGR"],
                    "fields": ["revenue", "revenue_yoy", "cagr"]
                },
                {
                    "title": "Consumer Market(Apparel)",
                    "r_type": "90000000",
                    "i_type": "apparel",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Footwear)",
                    "r_type": "11000000",
                    "i_type": "footwear",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Eyewear)",
                    "r_type": "12020000",
                    "i_type": "sunglasses",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Accessories)",
                    "r_type": "13000000",
                    "i_type": "accessories",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Luxury Goods)",
                    "r_type": "21000000",
                    "i_type": "luxury-goods",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                }
            ],
            "3C類": [
                {
                    "title": "Consumer Market",
                    "r_type": "15000000",
                    "i_type": "consumer-electronics",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Digital Market",
                    "r_type": "251",
                    "i_type": "consumer-electronics",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                },
                {
                    "title": "Consumer Market（Digital Camera）",
                    "r_type": "15010400",
                    "i_type": "digital-cameras",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market（mobile phones）",
                    "r_type": "15020100",
                    "i_type": "mobile-phones",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "（Laptops & Tablets）",
                    "r_type": "251",
                    "i_type": "consumer-electronics",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                }
            ],
            "家電類": [
                {
                    "title": "Consumer Market",
                    "r_type": "16000000",
                    "i_type": "household-appliances",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Digital Market",
                    "r_type": "256",
                    "i_type": "household-appliances",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                },
                {
                    "title": "Consumer Market（Vacuum Cleaners）",
                    "r_type": "16020100",
                    "i_type": "vacuum%25C2%25A0cleaners",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market（Small Kittchen appliance）",
                    "r_type": "16020200",
                    "i_type": "small-kitchen-appliances",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market（Coffee Machine）",
                    "r_type": "16021000",
                    "i_type": "coffee-machines",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                }
            ],
            "美妝類": [
                {
                    "title": "Consumer Market(Beauty & Personal Care)",
                    "r_type": "70000000",
                    "i_type": "beauty-personal-care",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Digital Market",
                    "r_type": "254",
                    "i_type": "personal-care",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                }
            ],
            "美妝次分類": [
                {
                    "title": "Consumer Market(cosmetics)",
                    "r_type": "70010000",
                    "i_type": "cosmetics",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Face Skincare)",
                    "r_type": "70020100",
                    "i_type": "face",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Body Skincare)",
                    "r_type": "70020200",
                    "i_type": "body",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(personal - hair care)",
                    "r_type": "70040000",
                    "i_type": "hair-care",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(personal - oral care)",
                    "r_type": "70060000",
                    "i_type": "oral-care",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(personal - fragrances)",
                    "r_type": "70050000",
                    "i_type": "fragrances",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                }
            ],
            "其他": [
                {
                    "title": "Digital Market - Toys, Hobby & DIY",
                    "r_type": "248",
                    "i_type": "toys-hobby-diy",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                },
                {
                    "title": "Digital Market - Sports & Outdoor",
                    "r_type": "259",
                    "i_type": "sports-outdoor",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                },
                {
                    "title": "Digital Market - Hobby & Stationary",
                    "r_type": "260",
                    "i_type": "hobby-stationery",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                },
                {
                    "title": "Consumer Market - Home & Laundry Care",
                    "r_type": "60000000",
                    "i_type": "home-laundry-care",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Tissue & Hygiene Paper",
                    "r_type": "80000000",
                    "i_type": "tissue-hygiene-paper",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Toys & Hobby",
                    "r_type": "19000000",
                    "i_type": "toys-hobby",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                }
            ],
            "五金類": [
                {
                    "title": "Consumer Market(Lamps & Lighting)",
                    "r_type": "17060000",
                    "i_type": "lamps-lighting",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Consumer Market(Floor Covering)",
                    "r_type": "17060000",
                    "i_type": "lamps-lighting",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPC"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpc"]
                },
                {
                    "title": "Digital Market(DIY, Garden & Pets)",
                    "r_type": "357",
                    "i_type": "diy-garden-pets",
                    "labels": ["Revenue, mln USD", "yoy", "CAGR", "ARPU"],
                    "fields": ["revenue", "revenue_yoy", "cagr", "arpu"]
                }
            ]
        }
    }

    prefix_label = data["prefix_label"]
    areas = data["areas"]
    tab = data["tab"]

    sheet_index = -1
    wb = Workbook()
    for tab_title in tab.keys():
        print(tab_title)
        # create sheet
        sheet_index += 1
        ws = wb.create_sheet(tab_title, sheet_index)
        # ws.cell(row=4, column=2, value=10)
        # ws.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
        cols_index = 1
        for data_group_id in range(len(tab[tab_title])):
            data_group = tab[tab_title][data_group_id]
            row_id = 1

            # print('\t{}\t{}\t{}\t{}'.format(data_group['title'], data_group['r_type'], data_group['i_type'],
            #                                 data_group['fields']))

            if data_group_id == 0:
                cols = prefix_label + [data_group['title']] + data_group['labels']
                col_start = 1
            else:
                cols = [data_group['title']] + data_group['labels']

            for c in range(len(cols)):
                _ = ws.cell(row=row_id, column=col_start+c, value="{}".format(cols[c]))

            for area in areas:
                row_id += 1
                url = "https://www.statista.com/outlook/{}/{}/{}/{}".format(data_group['r_type'],
                                                                            area[-2],
                                                                            data_group['i_type'],
                                                                            area[-1])

                # _ = ws.cell(row=row_id, column=cols_count, value="{}".format(cols[c]))
                # print("\t\t{}".format(url))
                crawler = Crawler(url)
                d = crawler.data()
                if data_group_id == 0:
                    row_data = [area[0], area[1], url]
                else:
                    row_data = [url]

                for f in data_group['fields']:
                    row_data.append(d[f])
                    # c = "{}\t{}-{}".format(c, f, d[f])

                for c in range(len(row_data)):
                    _ = ws.cell(row=row_id, column=col_start+c, value="{}".format(row_data[c]))

                print(row_data)

                sl = random.choice(SLEEP_SECOND)
                print("*************** sleep {} second ***************".format(sl))
                # if row_id >= 5:
                #     break

                time.sleep(sl)
            col_start += len(cols)

    wb.save(filename="/tmp/statista.xlsx")


if __name__ == "__main__":
    # app.run(threaded=True)
    run_crawler()
